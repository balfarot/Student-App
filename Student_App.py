# CSUMB Student App

from tkinter import *
import tkinter as tk
from tkinter.ttk import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = tk.Tk()

root.title('CSUMB Student App')
root.geometry('280x250')
root.resizable(0, 0)
root.grid_columnconfigure(1, weight=1)
root.config(background="gray95")

picture = PhotoImage(file="csumb1.png")
label = Label(root, image=picture)
label.grid(row=3, column=1)

wb = Workbook()
wb = load_workbook('Book1.xlsx')
ws = wb.active

column_a = ws['A']


def get_a():
    list = ''
    for cell in column_a:
        list = f'{list + str(cell.value)}\n'

        la.config(text=list)


ba = tk.Button(root, text="Select Student Names", font=("Arial", 15), bg="white", command=get_a)
ba.grid(row=2, column=1)

la = tk.Label(root, bg="gray95", font=("Arial", 10), text="")
la.grid(row=3, column=1)

tk.mainloop()
